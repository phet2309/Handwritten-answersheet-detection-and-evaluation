import tkinter as tk
from tkinter import *
import subprocess as sb_p
from xlrd import *
import openpyxl
import pandas as pd
import win32com.client
from setCamera2 import get_answer
import sys
import PIL.Image as Image
import numpy as np
import pytesseract
import cv2
import re
import array as ar
import os
# import the time module
import time



def verify(res,id):
    number = 0
    for i in range(2):
        for j in res:

            if j == str(id):
                number = 1

    return number

def set_exam(root,frame1):
    path_q='questions.txt'
    path_k='ans.xlsx'
    path_m='Marksheet.xlsx'

    for widget in frame1.winfo_children():
        widget.destroy()

    Label(frame1, text="Set Examination", font=('Helvetica', 25, 'bold')).grid(row = 0, column = 1, rowspan=1)
    Label(frame1, text="").grid(row = 1,column = 0)
    #Add Questions
    set_q = Button(frame1, text="Add Questions", width=15, command = lambda: os.system(path_q))

    #Set Keywords
    set_k = Button(frame1, text="Set Keywords", width=15, command = lambda: os.system(path_k))

    #Open Marksheet
    set_m = Button(frame1, text="Manage Marksheet", width=15, command = lambda: os.system(path_m))

    Label(frame1, text="").grid(row = 2,column = 0)
    Label(frame1, text="").grid(row = 4,column = 0)
    Label(frame1, text="").grid(row = 6,column = 0)
    set_q.grid(row = 3, column = 1, columnspan = 2)
    set_k.grid(row = 5, column = 1, columnspan = 2)
    set_m.grid(row = 7, column = 1, columnspan = 2)

    frame1.pack()
    root.mainloop()


def Evaluator(root, frame1):
    root.title("Set Examination")
    for widget in frame1.winfo_children():
        widget.destroy()
    Label(frame1, text="Enter Valid ID and password", font=('Helvetica', 24, 'bold')).grid(row=0, column=2, rowspan=1)
    Label(frame1, text="").grid(row=1, column=0)
    Label(frame1, text="ID:      ", anchor="e", justify=LEFT).grid(row=2, column=0)
    Label(frame1, text="Password:       ", anchor="e", justify=LEFT).grid(row=3, column=0)

    ID = tk.StringVar()
    password = tk.StringVar()

    e1 = Entry(frame1, textvariable=ID)
    e1.grid(row=2, column=2)
    e2 = Entry(frame1, textvariable=password)
    e2.grid(row=3, column=2)

    checking = Button(frame1, text="NEXT", width=15, command=lambda: eval_check(root, frame1, str(ID.get()), str(password.get())))
    Label(frame1, text="").grid(row=4, column=0)
    checking.grid(row=5, column=3, columnspan=2)
    frame1.pack()
    root.mainloop()


def eval_check(root, frame1, id, passw):
    if passw == "admin" and id == "admin":
        set_exam(root, frame1)
    else:
        Evaluator(root, frame1)


def Student(root, frame1):
    root.title("Start Examination")
    for widget in frame1.winfo_children():
        widget.destroy()
    Label(frame1, text="Enter Valid ID and your name", font=('Helvetica', 24, 'bold')).grid(row=0, column=2, rowspan=1)
    Label(frame1, text="").grid(row=1, column=0)
    Label(frame1, text="Student ID:      ", anchor="e", justify=LEFT).grid(row=2, column=0)
    Label(frame1, text="Name:       ", anchor="e", justify=LEFT).grid(row=3, column=0)

    s_ID = tk.StringVar()
    name = tk.StringVar()

    e1 = Entry(frame1, textvariable=s_ID)
    e1.grid(row=2, column=2)
    e2 = Entry(frame1, textvariable=name)
    e2.grid(row=3, column=2)

    path = 'questions.txt'
    checking = Button(frame1, text="START", width=15,
                      command=lambda: Q_paper(s_ID.get(), name.get(), root, frame1, load_questions(path)))
    Label(frame1, text="").grid(row=4, column=0)
    checking.grid(row=5, column=3, columnspan=2)
    frame1.pack()
    root.mainloop()


def check(id,qno):
    name = "C:/Users/SATYAPRAKASH/PycharmProjects/Test/AnswerSheet/"+str(id)+'_Q'+str(qno)+'.txt'
    file = open(name , 'r')
    test = file.read().replace("\n", " ")
    file.close()

    xlApp = win32com.client.Dispatch("Excel.Application")
    filename,password = r"C:/Users/SATYAPRAKASH/PycharmProjects/Test/ans.xlsx", '1234'
    xlwb = xlApp.Workbooks.Open(filename, False, True, None, password)
   # xlwb.Sheets().Select

    res = re.findall(r'\w+', test)
    flag = verify(res,id)
    if(flag == 0):
        print("Not Valid")
    number = 0

    k = 1
    while True :
        c=xlwb.Sheets(qno).Cells(k,1).Value
        print("form Sheet    :   " + str(c))
        if c is None:
            break
        k=k+1

    for i in range(k-2):
        for j in res:

            if j.lower() == xlwb.Sheets(qno).Cells(i+2,1).Value.lower():
                print(xlwb.Sheets(qno).Cells(i+2,2).Value)
                number = number+xlwb.Sheets(qno).Cells(i+2,2).Value

    print(number)
    wrkbk = openpyxl.load_workbook('Marksheet.xlsx')
    sh = wrkbk.active
    i = 1
    while True :
        c=sh.cell(row=i,column=1)
        if c.value == str(id):
            break
        i=i+1
    writer = pd.ExcelWriter('Marksheet.xlsx', engine='openpyxl')
    df = pd.read_excel('Marksheet.xlsx', header=None)
    df1 = pd.DataFrame([[number]])
    df.to_excel(writer, header=None, index=False)
    df1.to_excel(writer, header=None, index=False,
             startcol=qno+1,startrow=i-1)
    writer.save()

    xlwb.Close(SaveChanges=0)
    xlApp.Quit()



def load_questions(path):
    f = open(path, 'r')
    s=f.read()
    a=s.split('\n')
    return a


def Q_paper(s_ID,name,root,frame1,questions):

    for frame in root.winfo_children():
        for widget in frame.winfo_children():
            widget.destroy()

    wrkbk = openpyxl.load_workbook('Marksheet.xlsx')
    sh = wrkbk.active
    i = 1
    while True :
        c=sh.cell(row=i,column=1)
        if c.value is None:
            break
        i=i+1
    writer = pd.ExcelWriter('Marksheet.xlsx', engine='openpyxl')
    df = pd.read_excel('Marksheet.xlsx', header=None)
    df1 = pd.DataFrame([[s_ID  , name]])
    df.to_excel(writer, header=None, index=False)
    df1.to_excel(writer, header=None, index=False,
             startcol=0,startrow=i-1)
    writer.save()

    n = len(questions)
    q = []

    for j in range(0, n):

        Label(frame1, text="Q-" + str(j+1) + " :" + str(questions[j]) , font=('Helvetica', 24, 'bold')).grid(row = 2+2*j,column = 0)
        Label(frame1, text="").grid(row = 3+2*j,column = 0)

        q.append(None)
        q[j] = Button(frame1, text="START WRITING", width=20, command = lambda k=j :action(q,k,s_ID))
        q[j].grid(row = 2+2*j, column = 3, columnspan = 2)

    quit = Button(frame1, text="FINISH", width=20, command = lambda: root.quit())
    Label(frame1, text="").grid(row = 18,column = 0)
    quit.grid(row = 20, column = 0, columnspan = 2)

def action(q, j, s_ID):
    q[j].configure(state=DISABLED)
    get_answer(s_ID , j+1)
    #sb_p.call('start python setCamera.py', shell=True)
    #file_create(s_ID,j,string)
    check(s_ID,j+1)


# def file_create(s_ID,qno,string):
#     f = open(str(s_ID)+"_Q"+str(qno)+".txt", "w")
#     f.write(string)



def Home(root, frame1, frame2):

    for frame in root.winfo_children():
        for widget in frame.winfo_children():
            widget.destroy()

    Button(frame2, text="Home", command = lambda: Home(root, frame1, frame2)).grid(row=0,column=0)
    Label(frame2, text="                                                                         ").grid(row = 0,column = 1)
    Label(frame2, text="                                                                         ").grid(row = 0,column = 2)
    Label(frame2, text="         ").grid(row = 1,column = 1)
    frame2.pack(side=TOP)

    Label(frame1, text="Home", font=('Helvetica', 25, 'bold')).grid(row = 0, column = 1, rowspan=1)
    Label(frame1, text="").grid(row = 1,column = 0)
    #Evaluator Login
    eval = Button(frame1, text="Evaluator Login", width=15, command = lambda: Evaluator(root, frame1))

    #Student Login
    student = Button(frame1, text="Student Login", width=15, command = lambda: Student(root, frame1))

    Label(frame1, text="").grid(row = 2,column = 0)
    Label(frame1, text="").grid(row = 4,column = 0)
    eval.grid(row = 3, column = 1, columnspan = 2)
    student.grid(row = 5, column = 1, columnspan = 2)

    frame1.pack()
    root.mainloop()


root = Tk()
root.geometry('2000x2000')
root.title("Welcome...")
frame1 = Frame(root)
frame2 = Frame(root)
Home(root, frame1, frame2)










